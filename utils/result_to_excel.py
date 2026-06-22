import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image

import cv2
import numpy as np

def insert_images_to_excel(base_dir, mask_dir, inpaint_dir, output_excel):
    # 創建新的 Excel 檔案
    wb = Workbook()
    ws = wb.active

    # 支援的圖片格式
    supported_extensions = ('.png', '.jpg', '.jpeg', '.bmp')

    # 讀取 base_images 和 mask_images
    base_images = [f for f in os.listdir(base_dir) if f.endswith(supported_extensions)]
    mask_images = [f for f in os.listdir(mask_dir) if f.endswith(supported_extensions)]

    # 初始化黑色影像計數字典
    num_black_mask = {i: 0 for i in range(len(mask_images))}
    num_black_base = {j: 0 for j in range(len(base_images))}

    # 設定所有列寬 (A, B, C, ...)
    max_cols = len(base_images) + 1  # A 列 + base images 列
    for col_idx in range(1, max_cols + 1):  
        col_letter = chr(64 + col_idx)  # Excel: 1->A, 2->B, 3->C, ...
        ws.column_dimensions[col_letter].width = 15

    # 設定所有列高
    max_rows = len(mask_images) + 1  # 第一列是 base images，之後是 mask images
    for row_idx in range(1, max_rows + 1):
        ws.row_dimensions[row_idx].height = 80

    # 插入 base_images 到第一列 (從 B1 開始)
    for j, base_img in enumerate(base_images):
        img_path = os.path.join(base_dir, base_img)
        cell = f"{chr(66 + j)}1"  # B1, C1, D1...
        
        # 插入圖片
        img = Image(img_path)
        img.width = 100
        img.height = 100
        ws.add_image(img, cell)

    # 插入 mask_images 到第一欄 (從 A2 開始)
    for i, mask_img in enumerate(mask_images):
        img_path = os.path.join(mask_dir, mask_img)
        cell = f"A{2 + i}"  # A2, A3, ...
        
        # 插入圖片
        img = Image(img_path)
        img.width = 100
        img.height = 100
        ws.add_image(img, cell)

    # 插入 inpaint_{i}_{j}.png 到對應的交集 (B2 開始)
    for i, mask_img in enumerate(mask_images):
        for j, base_img in enumerate(base_images):
            inpaint_filename = f"inpaint_{i}_{j+1}.png"
            inpaint_path = os.path.join(inpaint_dir, inpaint_filename)

            if os.path.exists(inpaint_path):  # 確保檔案存在
                cell = f"{chr(66 + j)}{2+ i}"  # B2, C2, D2...

                # 讀取影像，判斷是否為全黑
                inpaint_image = cv2.imread(inpaint_path, cv2.IMREAD_GRAYSCALE)
                if inpaint_image is not None and np.all(inpaint_image == 0):  # 全黑影像
                    num_black_mask[i] += 1
                    num_black_base[j] += 1
                
                # 插入圖片
                img = Image(inpaint_path)
                img.width = 100
                img.height = 100
                ws.add_image(img, cell)
    
    # 插入 num_black_mask[i] 統計數據 (放在 B+len(base_images) 列)
    stat_col_mask = chr(66 + len(base_images))  # B + len(base_images)
    for i, count in num_black_mask.items():
        ws[f"{stat_col_mask}{2 + i}"] = count

    # 插入 num_black_base[j] 統計數據 (放在 B+len(base_images)+2 列)
    stat_row_base = len(mask_images)
    for j, count in num_black_base.items():
        ws[f"{chr(66 + j)}{stat_row_base + 2}"] = count # chr(66) -> B

    # 儲存 Excel 檔案
    wb.save(output_excel)
    print(f"Excel 檔案已儲存: {output_excel}")

if __name__ == "__main__":
    base_dir = r"datasets\test\clean_images(10)"
    mask_dir = r"datasets\test\masks"
    inpaint_dir = r"output\exp2_b=4_256x256"
    output_excel = "result.xlsx"

    insert_images_to_excel(base_dir, mask_dir, inpaint_dir, output_excel)
